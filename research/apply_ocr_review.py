"""Validate and apply teacher OCR decisions without overwriting the raw corpus."""

from __future__ import annotations

import argparse
from collections import Counter
from datetime import date, datetime
from hashlib import sha256
import json
from pathlib import Path
import sys
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
BACKEND = ROOT / "backend"
if str(BACKEND) not in sys.path:
    sys.path.insert(0, str(BACKEND))
SCRIPTS = ROOT / "scripts"
if str(SCRIPTS) not in sys.path:
    sys.path.insert(0, str(SCRIPTS))

from app.services.history_corpus import (  # noqa: E402
    HistoryChunk,
    extract_entity_candidates,
    extract_time_expressions,
    write_jsonl,
)


VALID_STATUSES = {"pending", "approved_as_is", "corrected", "rejected"}
REQUIRED_HEADERS = [
    "chunk_id", "risk_priority", "risk_score", "risk_flags", "topic_id",
    "lesson_id", "lesson_title", "pdf_page", "book_page", "ocr_confidence",
    "noise_character_count", "noise_density", "original_review_reasons",
    "recommended_action", "page_image", "source_text", "review_status",
    "corrected_text", "reviewer_id", "review_date", "review_comment",
]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--workbook",
        type=Path,
        default=Path("outputs/cobraq_ocr_review/CobraQ_OCR_Review.xlsx"),
    )
    parser.add_argument(
        "--chunks",
        type=Path,
        default=Path("data/research/history12_kntt/chunks.jsonl"),
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=Path("data/research/history12_kntt/chunks_reviewed.jsonl"),
    )
    parser.add_argument(
        "--audit-output",
        type=Path,
        default=Path("data/research/history12_kntt/ocr_review_audit.json"),
    )
    parser.add_argument(
        "--apply",
        action="store_true",
        help="Write reviewed outputs. Without this flag, only validate and report progress.",
    )
    parser.add_argument(
        "--reindex",
        action="store_true",
        help="Replace this book's vectors with the reviewed corpus after applying.",
    )
    parser.add_argument(
        "--corpus-dir",
        type=Path,
        default=Path("data/research/history12_kntt"),
    )
    parser.add_argument("--collection", default="cobraq_history12_kntt_v1")
    parser.add_argument(
        "--embedding-model", default="intfloat/multilingual-e5-small"
    )
    return parser.parse_args()


def file_sha256(path: Path) -> str:
    digest = sha256()
    with path.open("rb") as stream:
        for block in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def normalize_date(value: Any) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value or "").strip()


def load_corpus(path: Path) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    with path.open(encoding="utf-8") as stream:
        for line in stream:
            if line.strip():
                row = json.loads(line)
                HistoryChunk(**row)
                rows.append(row)
    return rows


def load_review_rows(path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=False)
    if "DUYET_OCR" not in workbook.sheetnames:
        raise ValueError("Workbook does not contain the DUYET_OCR sheet")
    sheet = workbook["DUYET_OCR"]
    headers = [sheet.cell(row=4, column=index).value for index in range(1, 22)]
    if headers != REQUIRED_HEADERS:
        raise ValueError("DUYET_OCR headers were changed; refusing to import")

    rows: list[dict[str, Any]] = []
    for values in sheet.iter_rows(min_row=5, max_col=21, values_only=True):
        if not values[0]:
            continue
        row = dict(zip(headers, values))
        row["chunk_id"] = str(row["chunk_id"]).strip()
        row["source_text"] = str(row["source_text"] or "")
        row["review_status"] = str(row["review_status"] or "pending").strip()
        row["corrected_text"] = str(row["corrected_text"] or "").strip()
        row["reviewer_id"] = str(row["reviewer_id"] or "").strip()
        row["review_date"] = normalize_date(row["review_date"])
        row["review_comment"] = str(row["review_comment"] or "").strip()
        rows.append(row)
    return rows


def validate(
    corpus: list[dict[str, Any]], review_rows: list[dict[str, Any]]
) -> tuple[dict[str, dict[str, Any]], dict[str, Any]]:
    corpus_by_id = {row["chunk_id"]: row for row in corpus}
    expected = {row["chunk_id"] for row in corpus if row.get("review_required")}
    review_ids = [row["chunk_id"] for row in review_rows]
    duplicates = sorted(item for item, count in Counter(review_ids).items() if count > 1)
    if duplicates:
        raise ValueError(f"Duplicate chunk_id values: {duplicates[:5]}")
    actual = set(review_ids)
    if actual != expected:
        missing = sorted(expected - actual)
        extra = sorted(actual - expected)
        raise ValueError(
            f"Review coverage mismatch: missing={missing[:5]}, extra={extra[:5]}"
        )

    errors: list[str] = []
    by_id: dict[str, dict[str, Any]] = {}
    for row in review_rows:
        chunk_id = row["chunk_id"]
        source = corpus_by_id[chunk_id]
        status = row["review_status"]
        if row["source_text"] != source["text"]:
            errors.append(f"{chunk_id}: source_text was modified")
        if status not in VALID_STATUSES:
            errors.append(f"{chunk_id}: invalid review_status={status!r}")
        if status == "corrected" and not row["corrected_text"]:
            errors.append(f"{chunk_id}: corrected status requires corrected_text")
        if status == "approved_as_is" and row["corrected_text"]:
            errors.append(f"{chunk_id}: approved_as_is must not contain corrected_text")
        if status != "pending":
            if not row["reviewer_id"]:
                errors.append(f"{chunk_id}: reviewer_id is required")
            if not row["review_date"]:
                errors.append(f"{chunk_id}: review_date is required")
        by_id[chunk_id] = row
    if errors:
        raise ValueError("Invalid review workbook:\n- " + "\n- ".join(errors[:30]))

    status_counts = Counter(row["review_status"] for row in review_rows)
    summary = {
        "total": len(review_rows),
        "pending": status_counts["pending"],
        "approved_as_is": status_counts["approved_as_is"],
        "corrected": status_counts["corrected"],
        "rejected": status_counts["rejected"],
        "ready_to_apply": status_counts["pending"] == 0,
    }
    return by_id, summary


def apply_reviews(
    corpus: list[dict[str, Any]],
    reviews: dict[str, dict[str, Any]],
    *,
    workbook_hash: str,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    reviewed: list[dict[str, Any]] = []
    audit_rows: list[dict[str, Any]] = []
    for source in corpus:
        chunk_id = source["chunk_id"]
        if chunk_id not in reviews:
            source.setdefault("ocr_review_status", "not_required")
            reviewed.append(source)
            continue

        decision = reviews[chunk_id]
        status = decision["review_status"]
        old_text = source["text"]
        old_hash = sha256(old_text.encode("utf-8")).hexdigest()
        audit_row = {
            "chunk_id": chunk_id,
            "status": status,
            "reviewer_id": decision["reviewer_id"],
            "review_date": decision["review_date"],
            "review_comment": decision["review_comment"],
            "original_text_sha256": old_hash,
        }

        if status == "rejected":
            audit_row["included_in_reviewed_corpus"] = False
            audit_rows.append(audit_row)
            continue

        item = dict(source)
        if status == "corrected":
            item["text"] = decision["corrected_text"]
            item["time_expressions"] = extract_time_expressions(item["text"])
            item["entity_candidates"] = extract_entity_candidates(item["text"])
        item["review_required"] = False
        item["review_reasons"] = []
        item["ocr_review_status"] = status
        item["ocr_reviewer_id"] = decision["reviewer_id"]
        item["ocr_review_date"] = decision["review_date"]
        item["ocr_review_comment"] = decision["review_comment"]
        item["ocr_original_text_sha256"] = old_hash
        item["ocr_review_workbook_sha256"] = workbook_hash
        HistoryChunk(**item)
        reviewed.append(item)
        audit_row["included_in_reviewed_corpus"] = True
        audit_row["final_text_sha256"] = sha256(item["text"].encode("utf-8")).hexdigest()
        audit_rows.append(audit_row)
    return reviewed, audit_rows


def main() -> int:
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8")
    args = parse_args()
    if args.reindex and not args.apply:
        raise ValueError("--reindex requires --apply")
    for required in (args.workbook, args.chunks):
        if not required.exists():
            raise FileNotFoundError(required)

    corpus = load_corpus(args.chunks)
    review_rows = load_review_rows(args.workbook)
    reviews, summary = validate(corpus, review_rows)
    summary.update(
        {
            "workbook": str(args.workbook.resolve()),
            "workbook_sha256": file_sha256(args.workbook),
            "source_chunks": str(args.chunks.resolve()),
            "source_chunks_sha256": file_sha256(args.chunks),
        }
    )
    print(json.dumps(summary, ensure_ascii=False, indent=2))
    if not args.apply:
        return 0
    if not summary["ready_to_apply"]:
        raise ValueError(
            f"Cannot apply: {summary['pending']} review rows are still pending"
        )

    reviewed, audit_rows = apply_reviews(
        corpus,
        reviews,
        workbook_hash=summary["workbook_sha256"],
    )
    write_jsonl(reviewed, args.output)
    args.audit_output.parent.mkdir(parents=True, exist_ok=True)
    audit = {
        "schema_version": "1.0",
        **summary,
        "reviewed_chunk_count": len(reviewed),
        "audit_rows": audit_rows,
    }
    if args.reindex:
        from ingest_history12 import index_chunks  # noqa: PLC0415

        reviewed_chunks = [HistoryChunk(**item) for item in reviewed]
        audit["vector_store"] = index_chunks(
            reviewed_chunks,
            output_dir=args.corpus_dir,
            collection=args.collection,
            embedding_model=args.embedding_model,
        )
    args.audit_output.write_text(
        json.dumps(audit, ensure_ascii=False, indent=2),
        encoding="utf-8",
        newline="\n",
    )
    print(f"reviewed_corpus={args.output.resolve()}")
    print(f"audit_log={args.audit_output.resolve()}")
    if args.reindex:
        print(json.dumps(audit["vector_store"], ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

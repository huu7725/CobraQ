"""Validate a teacher-reviewed AQG workbook and export LoRA-ready JSONL splits."""

from __future__ import annotations

import argparse
from collections import Counter, defaultdict
from datetime import date, datetime
from hashlib import sha256
import json
from pathlib import Path
import sys
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
BACKEND = ROOT / "backend"
for path in (ROOT, BACKEND):
    if str(path) not in sys.path:
        sys.path.insert(0, str(path))

from app.services.question_schema import GeneratedQuestion  # noqa: E402
from research.build_aqg_candidates import content_hash, file_sha256, load_jsonl, normalize  # noqa: E402
from research.score_teacher_rubric import score_row  # noqa: E402


DEFAULT_WORKBOOK = ROOT / "outputs" / "cobraq_aqg_review" / "CobraQ_AQG_600_Review.xlsx"
DEFAULT_CANDIDATES = ROOT / "data" / "research" / "aqg_v1" / "candidates_pending.jsonl"
DEFAULT_CORPUS = ROOT / "data" / "research" / "history12_kntt" / "chunks_reviewed.jsonl"
DEFAULT_RUBRIC = ROOT / "research" / "rubrics" / "teacher_rubric.json"
DEFAULT_OUTPUT_DIR = ROOT / "data" / "research" / "aqg_v1" / "approved"

QUESTION_HEADERS = [
    "stt", "question_id", "candidate_sha256", "lesson_id", "lesson_title", "book_page",
    "question_type", "difficulty", "bloom_level", "generation_method", "source_chunk_id",
    "evidence_quote", "draft_stem", "draft_choice_A", "draft_choice_B", "draft_choice_C",
    "draft_choice_D", "draft_correct_answer", "draft_explanation", "teacher_stem",
    "teacher_choice_A", "teacher_choice_B", "teacher_choice_C", "teacher_choice_D",
    "teacher_correct_answer", "teacher_explanation",
]
REVIEW_HEADERS = [
    "stt", "question_id", "question_type", "factual_accuracy", "source_support",
    "unique_correct_answer", "clarity", "curriculum_alignment", "cognitive_alignment",
    "distractor_quality", "pedagogical_language", "difficulty_fit", "weighted_score",
    "recommended_decision", "teacher_decision", "factual_error_found", "editing_minutes",
    "reviewer_id", "review_date", "comment",
]
DECISION_MAP = {"Chấp nhận": "approve", "Cần sửa": "revise", "Loại bỏ": "reject"}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK)
    parser.add_argument("--candidates", type=Path, default=DEFAULT_CANDIDATES)
    parser.add_argument("--corpus", type=Path, default=DEFAULT_CORPUS)
    parser.add_argument("--rubric", type=Path, default=DEFAULT_RUBRIC)
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR)
    parser.add_argument("--apply", action="store_true")
    parser.add_argument(
        "--allow-partial",
        action="store_true",
        help="Export fewer than 600 approved rows for pipeline smoke tests only.",
    )
    return parser.parse_args()


def normalize_date(value: Any) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value or "").strip()


def _sheet_rows(workbook: Any, sheet_name: str, headers: list[str]) -> list[dict[str, Any]]:
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"Workbook does not contain the {sheet_name} sheet")
    sheet = workbook[sheet_name]
    actual = [sheet.cell(row=4, column=index).value for index in range(1, len(headers) + 1)]
    if actual != headers:
        raise ValueError(f"{sheet_name} headers were changed; refusing to import")
    rows = []
    for values in sheet.iter_rows(min_row=5, max_col=len(headers), values_only=True):
        if not values[1]:
            continue
        rows.append(dict(zip(headers, values)))
    return rows


def load_workbook_rows(path: Path) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    workbook = load_workbook(path, read_only=True, data_only=False)
    return (
        _sheet_rows(workbook, "CAU_HOI", QUESTION_HEADERS),
        _sheet_rows(workbook, "DANH_GIA", REVIEW_HEADERS),
    )


def _draft_fields(record: dict[str, Any]) -> dict[str, Any]:
    question = record["response"]["questions"][0]
    choices = {choice["label"]: choice["text"] for choice in question.get("choices", [])}
    return {
        "question_id": record["record_id"],
        "candidate_sha256": record["candidate_content_sha256"],
        "lesson_id": record["lesson_id"],
        "lesson_title": record["lesson_title"],
        "book_page": record["source_book_pages"][0],
        "question_type": record["question_type"],
        "difficulty": record["difficulty"],
        "bloom_level": record["bloom_level"],
        "generation_method": record["candidate_generator"],
        "source_chunk_id": record["source_chunk_ids"][0],
        "evidence_quote": record["evidence_quote"],
        "draft_stem": question["stem"],
        "draft_choice_A": choices.get("A", ""),
        "draft_choice_B": choices.get("B", ""),
        "draft_choice_C": choices.get("C", ""),
        "draft_choice_D": choices.get("D", ""),
        "draft_correct_answer": question["correct_answer"],
        "draft_explanation": question["explanation"],
    }


def _text(value: Any) -> str:
    return str(value or "").strip()


def _score_value(value: Any) -> str:
    if value is None or value == "":
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _final_question(record: dict[str, Any], row: dict[str, Any]) -> dict[str, Any]:
    original = dict(record["response"]["questions"][0])
    original["stem"] = _text(row["teacher_stem"]) or original["stem"]
    original["correct_answer"] = _text(row["teacher_correct_answer"]) or original["correct_answer"]
    original["explanation"] = _text(row["teacher_explanation"]) or original["explanation"]
    if original["question_type"] == "multiple_choice":
        original_choices = {choice["label"]: choice["text"] for choice in original["choices"]}
        original["choices"] = [
            {
                "label": label,
                "text": _text(row[f"teacher_choice_{label}"]) or original_choices[label],
            }
            for label in "ABCD"
        ]
    return GeneratedQuestion.model_validate(original).model_dump()


def validate_review(
    candidates: list[dict[str, Any]], question_rows: list[dict[str, Any]],
    review_rows: list[dict[str, Any]], corpus: list[dict[str, Any]], rubric: dict[str, Any],
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    candidate_map = {row["record_id"]: row for row in candidates}
    question_map = {_text(row["question_id"]): row for row in question_rows}
    review_map = {_text(row["question_id"]): row for row in review_rows}
    expected_ids = set(candidate_map)
    errors: list[str] = []
    if set(question_map) != expected_ids:
        errors.append("CAU_HOI does not cover the exact candidate ID set")
    if set(review_map) != expected_ids:
        errors.append("DANH_GIA does not cover the exact candidate ID set")
    if len(question_map) != len(question_rows) or len(review_map) != len(review_rows):
        errors.append("Duplicate question_id values were found")
    chunk_map = {row["chunk_id"]: row for row in corpus}
    approved: list[dict[str, Any]] = []
    progress = Counter()

    for question_id, candidate in candidate_map.items():
        qrow = question_map.get(question_id)
        rrow = review_map.get(question_id)
        if qrow is None or rrow is None:
            continue
        for field, expected in _draft_fields(candidate).items():
            actual = qrow.get(field)
            if isinstance(expected, int):
                try:
                    actual = int(actual)
                except (TypeError, ValueError):
                    pass
            else:
                actual = _text(actual)
            if actual != expected:
                errors.append(f"{question_id}: protected field {field} was changed")
                break

        decision_label = _text(rrow["teacher_decision"])
        progress[decision_label or "Chưa chấm"] += 1
        if not decision_label or decision_label == "Chưa chấm":
            continue
        if decision_label not in DECISION_MAP:
            errors.append(f"{question_id}: invalid teacher_decision={decision_label!r}")
            continue
        reviewer_id = _text(rrow["reviewer_id"])
        review_date = normalize_date(rrow["review_date"])
        if not reviewer_id or not review_date:
            errors.append(f"{question_id}: reviewed rows require reviewer_id and review_date")
            continue
        score_input = {
            "question_id": question_id,
            "question_type": candidate["question_type"],
            **{criterion["id"]: _score_value(rrow[criterion["id"]]) for criterion in rubric["criteria"]},
            "factual_error_found": "true" if _text(rrow["factual_error_found"]) == "Có" else "false",
            "decision": DECISION_MAP[decision_label],
        }
        try:
            scored = score_row(score_input, rubric)
        except Exception as error:
            errors.append(f"{question_id}: rubric error: {error}")
            continue
        if decision_label != "Chấp nhận":
            continue
        if scored["computed_decision"] != "approve":
            errors.append(
                f"{question_id}: marked Chấp nhận but rubric computes {scored['computed_decision']}"
            )
            continue
        if _text(rrow["factual_error_found"]) != "Không":
            errors.append(f"{question_id}: accepted item must have factual_error_found=Không")
            continue
        try:
            question = _final_question(candidate, qrow)
        except Exception as error:
            errors.append(f"{question_id}: final question schema error: {error}")
            continue
        citation = question["citations"][0]
        source = chunk_map.get(citation["chunk_id"])
        if source is None or normalize(citation["quote"]) not in normalize(source["text"]):
            errors.append(f"{question_id}: final citation does not match reviewed corpus")
            continue
        approved_record = {
            **candidate,
            "response": {"questions": [question]},
            "review_status": "teacher_approved",
            "reviewer_id": reviewer_id,
            "review_date": review_date,
            "comment": _text(rrow["comment"]),
            "teacher_rubric": {
                "scores": {criterion["id"]: int(score_input[criterion["id"]]) for criterion in rubric["criteria"]
                           if score_input[criterion["id"]]},
                "weighted_score": scored["weighted_score"],
                "decision": "approve",
                "editing_minutes": float(rrow["editing_minutes"] or 0),
                "factual_error_found": False,
            },
        }
        approved_record["approved_content_sha256"] = content_hash({
            "response": approved_record["response"],
            "reviewer_id": reviewer_id,
            "review_date": review_date,
        })
        approved.append(approved_record)
    if errors:
        raise ValueError("Invalid AQG review workbook:\n- " + "\n- ".join(errors[:50]))
    summary = {
        "total_candidates": len(candidates),
        "review_progress": dict(progress),
        "teacher_approved": len(approved),
        "remaining_or_rejected": len(candidates) - len(approved),
        "ready_for_lora": len(approved) >= 600,
    }
    return approved, summary


def assign_splits(records: list[dict[str, Any]]) -> dict[str, list[dict[str, Any]]]:
    """Assign whole source chunks near an 80/10/10 lesson-stratified target."""
    groups: dict[tuple[str, str], list[dict[str, Any]]] = defaultdict(list)
    for record in records:
        groups[(record["lesson_id"], record["source_chunk_ids"][0])].append(record)
    lesson_ids = sorted({key[0] for key in groups})
    lesson_totals = {
        lesson_id: sum(len(items) for (lesson, _), items in groups.items() if lesson == lesson_id)
        for lesson_id in lesson_ids
    }
    global_holdout_target = round(len(records) * 0.1)

    def allocate_lesson_targets(salt: str) -> dict[str, int]:
        targets = {lesson_id: int(lesson_totals[lesson_id] * 0.1) for lesson_id in lesson_ids}
        remaining = global_holdout_target - sum(targets.values())
        order = sorted(
            lesson_ids,
            key=lambda lesson_id: (
                -(lesson_totals[lesson_id] * 0.1 - targets[lesson_id]),
                sha256(f"42:{salt}:{lesson_id}".encode()).hexdigest(),
            ),
        )
        for lesson_id in order[:max(0, remaining)]:
            targets[lesson_id] += 1
        return targets

    validation_targets = allocate_lesson_targets("validation")
    test_targets = allocate_lesson_targets("test")
    lesson_options: dict[str, list[dict[str, Any]]] = {}
    for lesson_id in lesson_ids:
        lesson_groups = [items for (lesson, _), items in groups.items() if lesson == lesson_id]
        lesson_groups.sort(
            key=lambda items: sha256(f"42:{items[0]['source_chunk_ids'][0]}".encode()).hexdigest()
        )
        total = lesson_totals[lesson_id]
        mcq_ratio = sum(
            item["question_type"] == "multiple_choice"
            for items in lesson_groups for item in items
        ) / total
        max_holdout = max(
            10,
            validation_targets[lesson_id] + 6,
            test_targets[lesson_id] + 6,
        )
        # State: validation count, test count, validation MCQ, test MCQ.
        states: dict[tuple[int, int, int, int], tuple[int, ...]] = {(0, 0, 0, 0): ()}
        for items in lesson_groups:
            size = len(items)
            mcq = sum(item["question_type"] == "multiple_choice" for item in items)
            updated: dict[tuple[int, int, int, int], tuple[int, ...]] = {}
            for state, assignment in states.items():
                choices = [
                    (state, assignment + (0,)),
                    ((state[0] + size, state[1], state[2] + mcq, state[3]), assignment + (1,)),
                    ((state[0], state[1] + size, state[2], state[3] + mcq), assignment + (2,)),
                ]
                for candidate_state, candidate_assignment in choices:
                    if candidate_state[0] > max_holdout or candidate_state[1] > max_holdout:
                        continue
                    current = updated.get(candidate_state)
                    if current is None or candidate_assignment < current:
                        updated[candidate_state] = candidate_assignment
            states = updated

        best_by_counts: dict[tuple[int, int], dict[str, Any]] = {}
        for (validation_count, test_count, validation_mcq, test_mcq), assignment in states.items():
            train_count = total - validation_count - test_count
            if train_count <= 0:
                continue
            if validation_targets[lesson_id] and validation_count == 0:
                continue
            if test_targets[lesson_id] and test_count == 0:
                continue
            count_cost = (
                abs(validation_count - validation_targets[lesson_id])
                + abs(test_count - test_targets[lesson_id])
            ) * 8
            composition_cost = (
                abs(validation_mcq - validation_count * mcq_ratio)
                + abs(test_mcq - test_count * mcq_ratio)
            )
            option = {
                "validation_count": validation_count,
                "test_count": test_count,
                "cost": count_cost + composition_cost,
                "assignment": assignment,
                "groups": lesson_groups,
            }
            key = (validation_count, test_count)
            current = best_by_counts.get(key)
            if current is None or (option["cost"], assignment) < (current["cost"], current["assignment"]):
                best_by_counts[key] = option
        if not best_by_counts:
            raise ValueError(f"Could not construct leakage-safe split options for {lesson_id}")
        lesson_options[lesson_id] = list(best_by_counts.values())

    # Combine lesson-level options while enforcing the global 10% targets.
    combined: dict[tuple[int, int], tuple[float, list[dict[str, Any]]]] = {(0, 0): (0.0, [])}
    cap = global_holdout_target + 10
    for lesson_id in lesson_ids:
        updated: dict[tuple[int, int], tuple[float, list[dict[str, Any]]]] = {}
        for (validation_total, test_total), (cost, selections) in combined.items():
            for option in lesson_options[lesson_id]:
                key = (
                    validation_total + option["validation_count"],
                    test_total + option["test_count"],
                )
                if key[0] > cap or key[1] > cap:
                    continue
                candidate = (cost + option["cost"], selections + [option])
                current = updated.get(key)
                if current is None or candidate[0] < current[0]:
                    updated[key] = candidate
        combined = updated
    if not combined:
        raise ValueError("Could not construct a global leakage-safe split")
    selected_key, (_, selected_options) = min(
        combined.items(),
        key=lambda item: (
            abs(item[0][0] - global_holdout_target) + abs(item[0][1] - global_holdout_target),
            abs(item[0][0] - item[0][1]),
            item[1][0],
            item[0],
        ),
    )
    if selected_key != (global_holdout_target, global_holdout_target):
        print(
            "warning: exact 80/10/10 grouped split was infeasible; "
            f"using validation={selected_key[0]}, test={selected_key[1]}"
        )

    splits = {"train": [], "validation": [], "test": []}
    split_names = {0: "train", 1: "validation", 2: "test"}
    for option in selected_options:
        for items, split_code in zip(option["groups"], option["assignment"]):
            split = split_names[split_code]
            for item in items:
                output = dict(item)
                output["split"] = split
                splits[split].append(output)
    for values in splits.values():
        values.sort(key=lambda row: row["record_id"])
    return splits


def write_jsonl(records: list[dict[str, Any]], path: Path) -> None:
    with path.open("w", encoding="utf-8", newline="\n") as stream:
        for record in records:
            stream.write(json.dumps(record, ensure_ascii=False) + "\n")


def main() -> int:
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8")
    args = parse_args()
    for required in (args.workbook, args.candidates, args.corpus, args.rubric):
        if not required.exists():
            raise FileNotFoundError(required)
    candidates = load_jsonl(args.candidates)
    corpus = load_jsonl(args.corpus)
    rubric = json.loads(args.rubric.read_text(encoding="utf-8"))
    question_rows, review_rows = load_workbook_rows(args.workbook)
    approved, summary = validate_review(candidates, question_rows, review_rows, corpus, rubric)
    summary.update({
        "workbook": str(args.workbook.resolve()),
        "workbook_sha256": file_sha256(args.workbook),
        "candidates_sha256": file_sha256(args.candidates),
    })
    print(json.dumps(summary, ensure_ascii=False, indent=2))
    if not args.apply:
        return 0
    if not approved:
        raise ValueError("No teacher-approved records are available to export")
    if not summary["ready_for_lora"] and not args.allow_partial:
        raise ValueError(
            "Fewer than 600 records passed teacher review; use --allow-partial only for a smoke test"
        )
    args.output_dir.mkdir(parents=True, exist_ok=True)
    splits = assign_splits(approved)
    for split, records in splits.items():
        write_jsonl(records, args.output_dir / f"{split}.jsonl")
    candidate_map = {row["record_id"]: row for row in candidates}
    weighted_scores = [row["teacher_rubric"]["weighted_score"] for row in approved]
    editing_minutes = [row["teacher_rubric"]["editing_minutes"] for row in approved]
    changed_from_draft = sum(
        row["response"] != candidate_map[row["record_id"]]["response"]
        for row in approved
    )
    unique_score_patterns = {
        tuple(sorted(row["teacher_rubric"]["scores"].items()))
        for row in approved
    }
    audit_warnings = []
    if len(set(weighted_scores)) == 1:
        audit_warnings.append("uniform_weighted_rubric_scores")
    if changed_from_draft == 0:
        audit_warnings.append("no_teacher_response_edits")
    if len({row["reviewer_id"] for row in approved}) == 1:
        audit_warnings.append("single_reviewer_training_annotation")
    manifest = {
        "schema_version": "1.0",
        **summary,
        "split_policy": "optimized_80_10_10_by_lesson_grouped_by_source_chunk_id_seed_42",
        "split_counts": {name: len(rows) for name, rows in splits.items()},
        "split_lesson_counts": {
            name: dict(sorted(Counter(row["lesson_id"] for row in rows).items()))
            for name, rows in splits.items()
        },
        "split_question_type_counts": {
            name: dict(Counter(row["question_type"] for row in rows))
            for name, rows in splits.items()
        },
        "split_difficulty_counts": {
            name: dict(sorted(Counter(str(row["difficulty"]) for row in rows).items()))
            for name, rows in splits.items()
        },
        "split_source_chunk_counts": {
            name: len({row["source_chunk_ids"][0] for row in rows})
            for name, rows in splits.items()
        },
        "review_audit_summary": {
            "reviewer_ids": sorted({row["reviewer_id"] for row in approved}),
            "review_dates": sorted({row["review_date"] for row in approved}),
            "responses_changed_from_draft": changed_from_draft,
            "responses_unchanged": len(approved) - changed_from_draft,
            "weighted_score_min": min(weighted_scores),
            "weighted_score_max": max(weighted_scores),
            "weighted_score_mean": round(sum(weighted_scores) / len(weighted_scores), 4),
            "unique_score_patterns": len(unique_score_patterns),
            "editing_minutes_total": round(sum(editing_minutes), 2),
            "editing_minutes_mean": round(sum(editing_minutes) / len(editing_minutes), 4),
            "warnings": audit_warnings,
            "interpretation": (
                "Training-data approval only; do not use these annotations as the blinded "
                "multi-rater C0-C3 effectiveness evaluation."
            ),
        },
        "split_hashes": {
            name: file_sha256(args.output_dir / f"{name}.jsonl") for name in splits
        },
    }
    (args.output_dir / "approved_manifest.json").write_text(
        json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8", newline="\n"
    )
    print(f"approved_dataset={args.output_dir.resolve()}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
